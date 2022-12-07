import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('produceSalesMini.xlsx')
sheet = wb['Sheet']
maxColLetter = get_column_letter(sheet.max_column)
maxRow = sheet.max_row
n = input()
m = input()

bottomCellRange = tuple(sheet['A'+str(int(n)+1):maxColLetter + str(maxRow)])
newCellRange = list(sheet['A'+str(int(n)+1+int(m)):maxColLetter + str(maxRow + int(m))])
coordinateValues = {}
for row in bottomCellRange:
  for cell in row:
    coordinateValues[cell.coordinate] = cell.value

col = 0
row = 0
for rows in newCellRange:
    for cell in rows:
        #celtrace = bottomCellRange[2][0].value
        #print(bottomCellRange[row][col].value)
        val = bottomCellRange[row][col].coordinate
        cell.value = coordinateValues[val]
        #print(cell.value)
        col += 1
    row +=1
    col = 0

deleteCellRange = tuple(sheet['A'+str(int(n)+1):maxColLetter + str(int(m)+int(n))])
for row in deleteCellRange:
    for cell in row:
        cell.value = None

wb.save('updatedProduceSalesMini.xlsx')



